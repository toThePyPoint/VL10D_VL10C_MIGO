import string

import openpyxl
import win32com.client
import pandas as pd
import pyperclip
import logging
import os
import io

from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime, timedelta
import numpy as np


def close_excel_file(file_name):
    # file_name = "mb52_table.xlsx"
    try:
        # Connect to the running Excel application
        excel = win32com.client.Dispatch("Excel.Application")

        # Iterate through open workbooks
        for workbook in excel.Workbooks:
            if workbook.FullName.endswith(file_name):  # Match the file name
                workbook.Save()  # Ensure the file is saved
                workbook.Close()  # Close the workbook
                print(f"{file_name} has been saved and closed.")
                break
        else:
            print(f"{file_name} not found in open Excel instances.")

        # Quit Excel if no other workbooks are open
        if excel.Workbooks.Count == 0:
            excel.Quit()

    except Exception as e:
        print(f"An error occurred: {e}")


def mb51_copy_data_from_excel_file(file_name="Arkusz w Basis (1)"):
    """
    Copies data from active sheet of open Excel file.
    :param file_name:
    :return:
    """
    # Connect to an open instance of Excel
    excel = win32com.client.GetActiveObject("Excel.Application")

    # Loop through open workbooks to find the correct one
    for wb in excel.Workbooks:
        if wb.Name == file_name:  # Match the file name
            sheet = wb.ActiveSheet  # Get the active sheet

            # Read data from the used range
            data = sheet.UsedRange.Value  # Get all data as a tuple of tuples

            # Convert data to a Pandas DataFrame
            df = pd.DataFrame(list(data))

            # Set the first row as column names (if the first row contains headers)
            df.columns = df.iloc[0]  # Assign first row as header
            df = df[1:].reset_index(drop=True)  # Remove the first row from data

            # Convert "Skł." to string and "Ilość" to integer
            df["Skł."] = df["Skł."].astype(str)
            df["Ilość"] = pd.to_numeric(df["Ilość"], errors='coerce').fillna(0).astype(int)

            # Convert DataFrame to clipboard-friendly format
            clipboard_data = df.to_csv(sep='\t', index=False)

            # Copy data to clipboard
            pyperclip.copy(clipboard_data)

            print("Data copied to clipboard!")
            break
    else:
        print("File not found in open Excel instances.")


def coois_copy_data_from_excel_file(file_name="Arkusz w Basis (1)"):
    """
    Copies data from active sheet of open Excel file.
    :param file_name:
    :return:
    """
    # Connect to an open instance of Excel
    excel = win32com.client.GetActiveObject("Excel.Application")

    # Loop through open workbooks to find the correct one
    for wb in excel.Workbooks:
        if wb.Name == file_name:  # Match the file name
            sheet = wb.ActiveSheet  # Get the active sheet

            # Read data from the used range
            data = sheet.UsedRange.Value  # Get all data as a tuple of tuples

            # Convert data to a Pandas DataFrame
            df = pd.DataFrame(list(data))

            # Set the first row as column names (if the first row contains headers)
            df.columns = df.iloc[0]  # Assign first row as header
            df = df[1:].reset_index(drop=True)  # Remove the first row from data

            # Convert "Skł." to string and "Ilość" to integer
            # df["Skł."] = df["Skł."].astype(str)
            # df["Ilość"] = pd.to_numeric(df["Ilość"], errors='coerce').fillna(0).astype(int)

            # Convert DataFrame to clipboard-friendly format
            clipboard_data = df.to_csv(sep='\t', index=False)

            # Copy data to clipboard
            pyperclip.copy(clipboard_data)

            print("Data copied to clipboard!")
            break
    else:
        print("File not found in open Excel instances.")


def copy_row_format(ws, source_row, target_row):
    """
    Copies border formatting from source_row to target_row in the given worksheet.

    :param ws: Worksheet object
    :param source_row: int: Row number to copy formatting from
    :param target_row: int: Row number to apply formatting to
    """
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)

        # Create a new Border object by copying the properties of the source cell
        if source_cell.border:
            new_border = Border(
                left=source_cell.border.left,
                right=source_cell.border.right,
                top=source_cell.border.top,
                bottom=source_cell.border.bottom
            )
            target_cell.border = new_border  # Apply new border

        # Copy Text Wrapping
        if source_cell.alignment:
            new_alignment = Alignment(
                wrap_text=source_cell.alignment.wrap_text  # Preserve wrapping
            )
            target_cell.alignment = new_alignment


def append_status_to_excel(status_file, status_dict, error_path, sheet_name):
    """
    Appends a new row to the "MRP_STOCKS" sheet in the given Excel file using the status_dict.

    :param sheet_name: sheet_name of excel status file
    :param error_path: path to error file
    :param status_file: str: Path to the Excel file
    :param status_dict: dict: Dictionary containing status messages
    """
    logging.basicConfig(
        filename=error_path,
        level=logging.ERROR,
        format="%(asctime)s - %(levelname)s - %(message)s",
    )
    try:
        # Load the existing workbook
        wb = load_workbook(status_file)

        # Select the "MRP_STOCKS" sheet
        if sheet_name not in wb.sheetnames:
            print(f"Error: Sheet '{sheet_name}' not found in the Excel file.")
            return

        ws = wb[sheet_name]

        # Get headers from the first row
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]

        # Find the first empty row within the framed area
        first_empty_row = None
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip headers
            if all(ws.cell(row=row, column=col).value in [None, ""] for col in range(1, ws.max_column + 1)):
                first_empty_row = row
                break

        # If no empty row is found, insert at the end
        if first_empty_row is None:
            first_empty_row = ws.max_row + 1

        # Insert new data at the first empty row
        ws.insert_rows(first_empty_row)

        # Copy border formatting from the row above
        if first_empty_row > 2:  # Ensure it's not the header row
            copy_row_format(ws, first_empty_row - 1, first_empty_row)

        # Add date/time in column A
        ws.cell(row=first_empty_row, column=1, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))

        # Fill in values based on dictionary keys matching headers
        for col, header in enumerate(headers[1:], start=2):  # Start from column 2 (B) as A is for timestamp
            ws.cell(row=first_empty_row, column=col, value=str(status_dict.get(header, "")))

        # Append the row and save the file
        wb.save(status_file)

        print("Row added successfully to STATUS FILE!")

    except Exception as e:
        logging.error("Error occurred", exc_info=True)
        print("Error occurred: ", e)
        print(f"Check {error_path} file for details")


def split_dataframe(df, chunk_size):
    """
    Splits a DataFrame into smaller chunks with a specified number of rows.

    :param df: Pandas DataFrame to split
    :param chunk_size: Number of rows per chunk
    :return: List of DataFrame chunks
    """
    return [df.iloc[i:i + chunk_size] for i in range(0, len(df), chunk_size)]


def get_last_n_working_days(n):
    """
    :param n: number of working days
    :return:
    """
    # Get today's date
    today = datetime.today().date()

    # Generate the last 15 working days (excluding weekends)
    working_days = [today - timedelta(days=i) for i in range(1, n*2) if
                    np.is_busday((today - timedelta(days=i)).strftime('%Y-%m-%d'))]

    # Keep only the last 15 working days
    last_15_working_days = working_days[:n]

    # Format the dates as 'dd.mm.yyyy'
    formatted_dates = [date.strftime('%d.%m.%Y') for date in last_15_working_days]

    return formatted_dates


def delete_file(filename):
    """
    Deletes the file specified by filename if it exists.

    Args:
        filename (str): The name of the file to delete.  Can be a relative or absolute path.

    Returns:
        bool: True if the file was successfully deleted, False if the file did not exist or there was an error.
    """
    try:
        if os.path.exists(filename):
            os.remove(filename)
            print(f"File '{filename}' deleted successfully.")
            return True
        else:
            print(f"File '{filename}' does not exist.")
            return False
    except Exception as e:
        print(f"Error deleting file '{filename}': {e}")
        return False


def vl10d_process_data(file_name_raw_data):
    # Wczytaj plik z pominięciem pustych wierszy i kolumn
    df_vl10d = pd.read_csv(file_name_raw_data, sep="\t", encoding='utf-16', skip_blank_lines=True)
    df_vl10d.dropna(how='all', inplace=True)  # usuwa całe puste wiersze
    # df_vl10d.dropna(axis=1, how='all', inplace=True)  # usuwa całe puste kolumny

    # Add goods_recepient_number column
    df_vl10d["goods_recepient_number"] = df_vl10d["Odb.mater."]

    new_columns_names = {
        "Unnamed: 3": "SAP_nr",
        "Unnamed: 5": "quantity",
        "Unnamed: 6": "quantity_2",
        "Unnamed: 7": "unit",
        "Unnamed: 9": "stock",
        "Unnamed: 12": "unit_2",
        "Unnamed: 13": "PrDst",
        "Unnamed: 15": "product_name",
        "Unnamed: 21": "goods_issue_date",
        "Unnamed: 23": "weight",
        "Unnamed: 25": "weight_unit",
        "Data utw.": "creation_date",
        "Odb.mater.": "doc_position",
        "Nazwa 1": "goods_recepient_name",
        "Autor": "author",
        "Dok.spraw.": "document_number",
        "BS": "sales_office"
    }

    df_vl10d.rename(columns=new_columns_names, inplace=True)

    # delete first row
    df_vl10d.drop(index=0, inplace=True)

    goods_recepient_number = None
    goods_recepient_name = None
    author = None
    doc_number = None
    sales_office = None

    for row in df_vl10d.iterrows():
        idx = row[0]
        # check if SAP_nr is NaN - if so, it's a delivery note's header row
        if pd.isna(row[1]["SAP_nr"]):
            # if so, set goods_recepient_number, goods_recepient_name and author
            goods_recepient_number = row[1]["goods_recepient_number"]
            goods_recepient_name = row[1]["goods_recepient_name"]
            author = row[1]["author"]
            doc_number = row[1]["document_number"]
            sales_office = row[1]["sales_office"]
        else:
            # if not, fill goods_recepient_number, goods_recepient_name and author columns
            # with the values from the last header row
            df_vl10d.at[idx, "goods_recepient_number"] = goods_recepient_number
            df_vl10d.at[idx, "goods_recepient_name"] = goods_recepient_name
            df_vl10d.at[idx, "author"] = author
            df_vl10d.at[idx, "document_number"] = doc_number
            df_vl10d.at[idx, "sales_office"] = sales_office

    # drop rows with NaN in SAP_nr column
    df_vl10d.dropna(subset=["SAP_nr"], inplace=True)
    # drop all empty columns, except "sales_office" column
    # df_vl10d.dropna(axis=1, how='all', inplace=True)  # usuwa całe puste kolumny
    cols_to_drop = [col for col in df_vl10d.columns if col != 'sales_office' and df_vl10d[col].isna().all()]
    df_vl10d.drop(columns=cols_to_drop, inplace=True)

    # drop columns that are not needed
    columns_to_drop = [
        "quantity_2",
        "RDok",
        "DSprz",
        "Trasa",
        "unit_2",
        "PrDst",
        "   Waga brutto",
        "JWg",
        "IncoT",
        "Inco. 2",
        "Incoterms 2",
        "weight",
        "creation_date",
        "weight_unit",
        "Zamówienie",
        "KDs",
    ]
    # Column names has different values. Sometimes they are shortened
    valid_columns_to_drop = [col for col in columns_to_drop if col in df_vl10d.columns]
    df_vl10d.drop(columns=valid_columns_to_drop, inplace=True)

    strings_to_filter_out_1 = ['ZRV', 'ZAR', 'ZRI', 'ZJA', 'ZRE', 'R4', 'R7', 'ZFA', 'R6', 'R8', 'Q4', 'R3', 'R2',
                               'Behang Screen', 'EFL', 'ABR', 'R5', 'ZIN', 'ERS', 'ASA', 'ASI', 'MDA', 'POS',
                               'ZRO', 'ZRS', 'EDH', 'EPA', 'EDL', 'EDZ', 'ED_', 'Ständer', 'Koszty transportu',
                               'EDQ', 'EDT', 'PALETTE', 'EDF', 'EA', 'EDS', 'EDW', 'EDG']
    strings_to_filter_out_2 = ["WROBELM", "KICIAM", "PLATINE", "MONTAZS100", "POLICHANCZUK", "WOZNIAKT"]
    # strings_to_filter_out_3 = ["103702"]
    strings_to_filter_out_4 = ["99"]
    strings_to_filter_out_5 = ['Artikel']

    # Use the .str.startswith() method on the specified column and negate the boolean mask
    df_filtered = df_vl10d[~df_vl10d['product_name'].str.startswith(tuple(strings_to_filter_out_1))]
    df_filtered = df_filtered[~df_filtered['author'].isin(strings_to_filter_out_2)]
    # df_filtered = df_filtered[~df_filtered['goods_recepient_number'].isin(strings_to_filter_out_3)]
    df_filtered = df_filtered[~df_filtered['SAP_nr'].str.startswith(tuple(strings_to_filter_out_4))]
    df_filtered = df_filtered[~df_filtered['product_name'].isin(strings_to_filter_out_5)]
    df_filtered = df_filtered[
        ~df_filtered['product_name'].str.contains(r'Ausstellarm.*kpl.*|Eckumlenkung.*kpl.*|Gasfeder.*kpl.*', case=False, na=False)]

    df_filtered.insert(loc=2, column='is_booking_req', value='n')

    # sort by goods issue date
    df_filtered['goods_issue_date'] = pd.to_datetime(df_filtered['goods_issue_date'], format='%d.%m.%Y').dt.date
    df_sorted = df_filtered.sort_values(by=['goods_issue_date', 'document_number'], ascending=[True, True])

    # ensure correct quantity data types
    df_sorted['quantity'] = df_sorted['quantity'].apply(lambda x: float(str(x).replace('.', '').replace(',', '.').strip()))
    df_sorted['stock'] = df_sorted['stock'].apply(lambda x: float(str(x).replace('.', '').replace(',', '.').strip()))

    return df_sorted


def run_excel_file_and_adjust_col_width(file_path):
    """
    Opens the Excel file using the operating system's default application.

    Args:
        file_path (str): The path to the Excel file to be opened.

    Returns:
        bool: True if the file was opened successfully, False otherwise.
    """
    try:
        # Check if the file exists
        if not os.path.exists(file_path):
            print(f"File does not exist: {file_path}")
            return False

        # Load the workbook to adjust column widths
        try:
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active

            # Adjust column widths for columns A to J
            for column_letter in string.ascii_uppercase:
                max_length = 0
                for cell in sheet[column_letter]:
                    try:
                        cell_value_length = len(str(cell.value))
                        if cell_value_length > max_length:
                            max_length = cell_value_length
                    except TypeError:
                        pass  # Handle cases where cell.value is None
                adjusted_width = (max_length + 2)  # Add some padding
                sheet.column_dimensions[column_letter].width = adjusted_width

            workbook.save(file_path)  # Save changes to the file
        except Exception as e:
            print(f"Error adjusting column widths: {e}")

        # Open the file using the default application
        if os.name == 'nt':  # For Windows
            os.startfile(file_path)
        else:
            print("Unsupported operating system.")
            return False
        return True

    except Exception as e:
        print(f"An error occurred while opening the file: {e}")
        return False


def copy_df_column_to_clipboard(df, column_name):
    """
    Copies the specified column from a pandas DataFrame to the clipboard using pyperclip.

    Args:
        df (pd.DataFrame): The pandas DataFrame containing the data.
        column_name (str): The name of the column to copy.

    Returns:
        bool: True if the data was copied successfully, False otherwise.
    """
    try:
        # Check if the column exists in the DataFrame
        if column_name not in df.columns:
            print(f"Error: Column '{column_name}' not found in DataFrame.")
            return False

        # # Select the column and convert it to a string format
        # column_data = df[column_name].astype(str)
        # column_string = column_data.to_string(header=True, index=False)

        # Select the column
        column_data = df[column_name]

        # Convert to string with tab separation
        output = io.StringIO()
        column_data.to_csv(output, sep='\t', header=False, index=False)
        column_string = output.getvalue()
        output.close()

        # Copy the string to the clipboard using pyperclip
        pyperclip.copy(column_string)
        return True
    except Exception as e:
        print(f"An error occurred: {e}")
        return False
